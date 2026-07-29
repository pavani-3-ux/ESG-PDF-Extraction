import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

INPUT_FILE = os.path.join(
    BASE_DIR,
    "output",
    "llm",
    "llm_generated_answers.xlsx"
)

OUTPUT_FOLDER = os.path.join(
    BASE_DIR,
    "output",
    "validation"
)

OUTPUT_FILE = os.path.join(
    OUTPUT_FOLDER,
    "llm_answer_validation_report.xlsx"
)

# ============================================================
# HEADER
# ============================================================

print("=" * 70)
print("LLM ANSWER VALIDATION")
print("=" * 70)

# ============================================================
# CHECK INPUT FILE
# ============================================================

if not os.path.exists(INPUT_FILE):
    print("\n❌ LLM answer file not found!")
    print(f"Expected location:\n{INPUT_FILE}")

    raise FileNotFoundError(
        f"File not found: {INPUT_FILE}"
    )

print("\n✅ LLM answer file found!")
print(f"Location: {INPUT_FILE}")

# ============================================================
# LOAD EXCEL FILE
# ============================================================

print("\nLoading LLM answers...")

df = pd.read_excel(
    INPUT_FILE
)

print("✅ LLM answers loaded successfully!")

print(f"Rows: {len(df)}")
print(f"Columns: {len(df.columns)}")

print("\nInput columns:")

for column in df.columns:
    print(f"• {column}")

# ============================================================
# COLUMN DETECTION
# ============================================================

def find_column(
    dataframe,
    candidates
):

    for candidate in candidates:

        if candidate in dataframe.columns:
            return candidate

    return None


query_column = find_column(
    df,
    [
        "query",
        "Query",
        "question",
        "Question"
    ]
)

context_column = find_column(
    df,
    [
        "context",
        "Context",
        "combined_context",
        "Combined Context"
    ]
)

answer_column = find_column(
    df,
    [
        "llm_answer",
        "LLM Answer",
        "answer",
        "Answer"
    ]
)

status_column = find_column(
    df,
    [
        "status",
        "Status"
    ]
)

error_column = find_column(
    df,
    [
        "error",
        "Error"
    ]
)

print("\n" + "=" * 70)
print("COLUMN DETECTION")
print("=" * 70)

print(f"Query column   : {query_column}")
print(f"Context column : {context_column}")
print(f"Answer column  : {answer_column}")
print(f"Status column  : {status_column}")
print(f"Error column   : {error_column}")

# ============================================================
# VALIDATE REQUIRED COLUMNS
# ============================================================

if query_column is None:

    raise ValueError(
        "Query column not found!"
    )

if context_column is None:

    raise ValueError(
        "Context column not found!"
    )

if answer_column is None:

    raise ValueError(
        "LLM answer column not found!"
    )

# ============================================================
# CREATE VALIDATION RESULTS
# ============================================================

validation_results = []

print("\n" + "=" * 70)
print("RUNNING AUTOMATIC VALIDATION")
print("=" * 70)

# ============================================================
# PROCESS EACH ANSWER
# ============================================================

for index, row in df.iterrows():

    query = str(
        row[query_column]
    ).strip()

    context = str(
        row[context_column]
    ).strip()

    answer = str(
        row[answer_column]
    ).strip()

    # --------------------------------------------------------
    # CHECK QUERY
    # --------------------------------------------------------

    query_present = (
        bool(query)
        and query.lower() != "nan"
    )

    # --------------------------------------------------------
    # CHECK CONTEXT
    # --------------------------------------------------------

    context_present = (
        bool(context)
        and context.lower() != "nan"
    )

    # --------------------------------------------------------
    # CHECK ANSWER
    # --------------------------------------------------------

    answer_present = (
        bool(answer)
        and answer.lower() != "nan"
    )

    # --------------------------------------------------------
    # ANSWER LENGTH
    # --------------------------------------------------------

    if answer_present:

        answer_length = len(
            answer
        )

    else:

        answer_length = 0

    # --------------------------------------------------------
    # CHECK "NOT AVAILABLE" RESPONSE
    # --------------------------------------------------------

    unavailable_phrases = [

        "information is not available",

        "not available in the provided context",

        "not provided in the context",

        "cannot be determined from the context",

        "not mentioned in the provided context",

        "not found in the provided context"

    ]

    information_not_available = any(

        phrase in answer.lower()

        for phrase in unavailable_phrases

    )

    # --------------------------------------------------------
    # CHECK ERROR
    # --------------------------------------------------------

    error_message = ""

    if error_column is not None:

        error_message = str(
            row[error_column]
        ).strip()

        if error_message.lower() == "nan":

            error_message = ""

    # --------------------------------------------------------
    # DETERMINE AUTOMATIC STATUS
    # --------------------------------------------------------

    if not query_present:

        automatic_status = "FAILED"

        validation_reason = (
            "Query is missing"
        )

    elif not context_present:

        automatic_status = "FAILED"

        validation_reason = (
            "Context is missing"
        )

    elif not answer_present:

        automatic_status = "FAILED"

        validation_reason = (
            "LLM answer is missing"
        )

    elif error_message:

        automatic_status = "FAILED"

        validation_reason = (
            "LLM generation error found"
        )

    elif information_not_available:

        automatic_status = "REVIEW"

        validation_reason = (
            "LLM reported information unavailable"
        )

    elif answer_length < 20:

        automatic_status = "REVIEW"

        validation_reason = (
            "Answer is very short"
        )

    else:

        automatic_status = "PASS"

        validation_reason = (
            "Basic automatic validation passed"
        )

    # --------------------------------------------------------
    # STORE RESULT
    # --------------------------------------------------------

    validation_results.append({

        "query":
            query,

        "context":
            context,

        "llm_answer":
            answer,

        "answer_length":
            answer_length,

        "query_present":
            "YES"
            if query_present
            else "NO",

        "context_present":
            "YES"
            if context_present
            else "NO",

        "answer_present":
            "YES"
            if answer_present
            else "NO",

        "information_not_available":
            "YES"
            if information_not_available
            else "NO",

        "automatic_status":
            automatic_status,

        "validation_reason":
            validation_reason,

        # ----------------------------------------------------
        # MANUAL VALIDATION COLUMNS
        # ----------------------------------------------------

        "answer_relevant":
            "",

        "answer_grounded":
            "",

        "answer_correct":
            "",

        "hallucination_detected":
            "",

        "source_supported":
            "",

        "review_comments":
            ""

    })

# ============================================================
# CREATE VALIDATION DATAFRAME
# ============================================================

validation_df = pd.DataFrame(
    validation_results
)

# ============================================================
# CREATE OUTPUT FOLDER
# ============================================================

os.makedirs(
    OUTPUT_FOLDER,
    exist_ok=True
)

# ============================================================
# CREATE SUMMARY
# ============================================================

total_answers = len(
    validation_df
)

passed = (
    validation_df[
        "automatic_status"
    ] == "PASS"
).sum()

failed = (
    validation_df[
        "automatic_status"
    ] == "FAILED"
).sum()

review = (
    validation_df[
        "automatic_status"
    ] == "REVIEW"
).sum()

summary_df = pd.DataFrame({

    "Metric": [

        "Total Answers",

        "Passed",

        "Failed",

        "Needs Review"

    ],

    "Count": [

        total_answers,

        passed,

        failed,

        review

    ]

})

# ============================================================
# SAVE EXCEL REPORT
# ============================================================

print("\n" + "=" * 70)
print("CREATING VALIDATION REPORT")
print("=" * 70)

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    validation_df.to_excel(

        writer,

        sheet_name="LLM Answer Validation",

        index=False

    )

    summary_df.to_excel(

        writer,

        sheet_name="Validation Summary",

        index=False

    )

# ============================================================
# FORMAT EXCEL FILE
# ============================================================

print("\nFormatting Excel report...")

workbook = load_workbook(
    OUTPUT_FILE
)

# ------------------------------------------------------------
# FORMAT VALIDATION SHEET
# ------------------------------------------------------------

validation_sheet = workbook[
    "LLM Answer Validation"
]

# Freeze first row

validation_sheet.freeze_panes = "A2"

# Header formatting

for cell in validation_sheet[1]:

    cell.font = Font(
        bold=True
    )

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

# Wrap text

for row in validation_sheet.iter_rows():

    for cell in row:

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

# Auto width

for column_cells in validation_sheet.columns:

    column_letter = get_column_letter(
        column_cells[0].column
    )

    max_length = 0

    for cell in column_cells:

        try:

            cell_length = len(
                str(cell.value)
            )

            if cell_length > max_length:

                max_length = cell_length

        except:

            pass

    validation_sheet.column_dimensions[
        column_letter
    ].width = min(
        max_length + 2,
        60
    )

# ------------------------------------------------------------
# FORMAT SUMMARY SHEET
# ------------------------------------------------------------

summary_sheet = workbook[
    "Validation Summary"
]

summary_sheet.freeze_panes = "A2"

for cell in summary_sheet[1]:

    cell.font = Font(
        bold=True
    )

    cell.alignment = Alignment(
        horizontal="center"
    )

for column_cells in summary_sheet.columns:

    column_letter = get_column_letter(
        column_cells[0].column
    )

    summary_sheet.column_dimensions[
        column_letter
    ].width = 25

# ============================================================
# SAVE FORMATTED WORKBOOK
# ============================================================

workbook.save(
    OUTPUT_FILE
)

# ============================================================
# FINAL OUTPUT
# ============================================================

print("\n" + "=" * 70)
print("LLM ANSWER VALIDATION COMPLETED")
print("=" * 70)

print(f"\nTotal answers: {total_answers}")

print(f"Passed: {passed}")

print(f"Failed: {failed}")

print(f"Needs review: {review}")

print("\nValidation report saved at:")

print(OUTPUT_FILE)

print("\n" + "=" * 70)

print("NEXT STEP")

print("=" * 70)

print(
    "\nOpen the Excel report and manually review:"
)

print(
    "1. answer_relevant"
)

print(
    "2. answer_grounded"
)

print(
    "3. answer_correct"
)

print(
    "4. hallucination_detected"
)

print(
    "5. source_supported"
)

print(
    "6. review_comments"
)

print(
    "\nAfter completing manual validation,"
)

print(
    "the next stage is Answer Quality Evaluation "
    "and Source Attribution."
)