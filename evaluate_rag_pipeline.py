import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

RETRIEVAL_FILE = os.path.join(
    BASE_DIR,
    "output",
    "retrieval",
    "retrieval_similarity_search_results.xlsx"
)

LLM_FILE = os.path.join(
    BASE_DIR,
    "output",
    "llm",
    "llm_generated_answers.xlsx"
)

OUTPUT_FOLDER = os.path.join(
    BASE_DIR,
    "output",
    "evaluation"
)

OUTPUT_FILE = os.path.join(
    OUTPUT_FOLDER,
    "rag_end_to_end_evaluation_report.xlsx"
)

# ============================================================
# HEADER
# ============================================================

print("=" * 70)
print("RAG END-TO-END EVALUATION")
print("=" * 70)

# ============================================================
# CHECK FILES
# ============================================================

print("\nChecking input files...")

if not os.path.exists(RETRIEVAL_FILE):

    raise FileNotFoundError(
        f"Retrieval file not found:\n{RETRIEVAL_FILE}"
    )

print("✅ Retrieval file found!")

if not os.path.exists(LLM_FILE):

    raise FileNotFoundError(
        f"LLM answer file not found:\n{LLM_FILE}"
    )

print("✅ LLM answer file found!")

# ============================================================
# LOAD FILES
# ============================================================

print("\nLoading retrieval results...")

retrieval_df = pd.read_excel(
    RETRIEVAL_FILE
)

print(
    f"Retrieval rows: {len(retrieval_df)}"
)

print("\nLoading LLM answers...")

llm_df = pd.read_excel(
    LLM_FILE
)

print(
    f"LLM answer rows: {len(llm_df)}"
)

# ============================================================
# DISPLAY COLUMNS
# ============================================================

print("\nRetrieval columns:")

for column in retrieval_df.columns:

    print(
        f"• {column}"
    )

print("\nLLM columns:")

for column in llm_df.columns:

    print(
        f"• {column}"
    )

# ============================================================
# NORMALIZE QUERY
# ============================================================

retrieval_df["query"] = (
    retrieval_df["query"]
    .fillna("")
    .astype(str)
    .str.strip()
)

llm_df["query"] = (
    llm_df["query"]
    .fillna("")
    .astype(str)
    .str.strip()
)

# ============================================================
# RETRIEVAL METRICS
# ============================================================

print("\n" + "=" * 70)
print("CALCULATING RETRIEVAL METRICS")
print("=" * 70)

retrieval_summary = []

for query in retrieval_df["query"].unique():

    query_data = retrieval_df[
        retrieval_df["query"] == query
    ].copy()

    total_results = len(
        query_data
    )

    # --------------------------------------------------------
    # TOP RESULT
    # --------------------------------------------------------

    top_result = (
        query_data
        .sort_values("rank")
        .iloc[0]
    )

    top_similarity = float(
        top_result[
            "similarity_score"
        ]
    )

    top_distance = float(
        top_result[
            "distance"
        ]
    )

    top_company = str(
        top_result[
            "company"
        ]
    )

    top_source = str(
        top_result[
            "source_file"
        ]
    )

    # --------------------------------------------------------
    # AVERAGE SIMILARITY
    # --------------------------------------------------------

    average_similarity = (
        query_data[
            "similarity_score"
        ]
        .mean()
    )

    # --------------------------------------------------------
    # CONTENT TYPE DISTRIBUTION
    # --------------------------------------------------------

    content_types = (
        query_data[
            "content_type"
        ]
        .fillna("unknown")
        .astype(str)
        .value_counts()
        .to_dict()
    )

    text_count = content_types.get(
        "text",
        0
    )

    table_count = content_types.get(
        "table",
        0
    )

    image_count = content_types.get(
        "image",
        0
    )

    # --------------------------------------------------------
    # RETRIEVAL QUALITY
    # --------------------------------------------------------

    if top_similarity >= 0.80:

        retrieval_quality = "Excellent"

    elif top_similarity >= 0.60:

        retrieval_quality = "Good"

    elif top_similarity >= 0.40:

        retrieval_quality = "Moderate"

    else:

        retrieval_quality = "Low"

    retrieval_summary.append({

        "query":
            query,

        "total_retrieved":
            total_results,

        "top_similarity_score":
            round(
                top_similarity,
                4
            ),

        "average_similarity_score":
            round(
                average_similarity,
                4
            ),

        "top_distance":
            round(
                top_distance,
                4
            ),

        "top_company":
            top_company,

        "top_source":
            top_source,

        "text_results":
            text_count,

        "table_results":
            table_count,

        "image_results":
            image_count,

        "retrieval_quality":
            retrieval_quality

    })

retrieval_metrics_df = pd.DataFrame(
    retrieval_summary
)

print(
    "✅ Retrieval metrics calculated!"
)

# ============================================================
# LLM METRICS
# ============================================================

print("\n" + "=" * 70)
print("CALCULATING LLM METRICS")
print("=" * 70)

llm_evaluation = []

for index, row in llm_df.iterrows():

    query = str(
        row["query"]
    ).strip()

    context = str(
        row["context"]
    ).strip()

    answer = str(
        row["llm_answer"]
    ).strip()

    status = str(
        row["status"]
    ).strip()

    error = str(
        row["error"]
    ).strip()

    # --------------------------------------------------------
    # CONTEXT CHECK
    # --------------------------------------------------------

    context_present = (

        bool(context)

        and

        context.lower() != "nan"

    )

    # --------------------------------------------------------
    # ANSWER CHECK
    # --------------------------------------------------------

    answer_present = (

        bool(answer)

        and

        answer.lower() != "nan"

    )

    # --------------------------------------------------------
    # ANSWER LENGTH
    # --------------------------------------------------------

    answer_length = len(
        answer
    )

    # --------------------------------------------------------
    # INFORMATION NOT AVAILABLE
    # --------------------------------------------------------

    unavailable_phrases = [

        "information is not available",

        "not available in the provided context",

        "not provided in the context",

        "cannot be determined from the context",

        "not mentioned in the provided context"

    ]

    information_not_available = any(

        phrase in answer.lower()

        for phrase in unavailable_phrases

    )

    # --------------------------------------------------------
    # AUTOMATIC LLM QUALITY
    # --------------------------------------------------------

    if status == "SUCCESS" and answer_present:

        if information_not_available:

            llm_quality = "Needs Review"

        elif answer_length >= 100:

            llm_quality = "Good"

        elif answer_length >= 30:

            llm_quality = "Moderate"

        else:

            llm_quality = "Low"

    else:

        llm_quality = "Failed"

    llm_evaluation.append({

        "query":
            query,

        "context_present":
            "YES"
            if context_present
            else "NO",

        "answer_present":
            "YES"
            if answer_present
            else "NO",

        "answer_length":
            answer_length,

        "information_not_available":
            "YES"
            if information_not_available
            else "NO",

        "llm_status":
            status,

        "llm_quality":
            llm_quality,

        "error":
            error,

        "llm_answer":
            answer

    })

llm_metrics_df = pd.DataFrame(
    llm_evaluation
)

print(
    "✅ LLM metrics calculated!"
)

# ============================================================
# MERGE RETRIEVAL + LLM
# ============================================================

print("\n" + "=" * 70)
print("CREATING END-TO-END EVALUATION")
print("=" * 70)

end_to_end_df = pd.merge(

    retrieval_metrics_df,

    llm_metrics_df,

    on="query",

    how="outer"

)

# ============================================================
# END-TO-END STATUS
# ============================================================

def calculate_end_to_end_status(row):

    retrieval_quality = str(
        row.get(
            "retrieval_quality",
            ""
        )
    )

    llm_quality = str(
        row.get(
            "llm_quality",
            ""
        )
    )

    if (

        retrieval_quality
        in [
            "Excellent",
            "Good"
        ]

        and

        llm_quality
        == "Good"

    ):

        return "PASS"

    elif (

        retrieval_quality
        in [
            "Excellent",
            "Good",
            "Moderate"
        ]

        and

        llm_quality
        in [
            "Good",
            "Moderate",
            "Needs Review"
        ]

    ):

        return "REVIEW"

    else:

        return "FAIL"

end_to_end_df[
    "end_to_end_status"
] = end_to_end_df.apply(

    calculate_end_to_end_status,

    axis=1

)

# ============================================================
# CREATE SUMMARY
# ============================================================

total_queries = len(
    end_to_end_df
)

passed_queries = (

    end_to_end_df[
        "end_to_end_status"
    ]

    == "PASS"

).sum()

review_queries = (

    end_to_end_df[
        "end_to_end_status"
    ]

    == "REVIEW"

).sum()

failed_queries = (

    end_to_end_df[
        "end_to_end_status"
    ]

    == "FAIL"

).sum()

average_top_similarity = (

    end_to_end_df[
        "top_similarity_score"
    ]

    .mean()

)

average_answer_length = (

    end_to_end_df[
        "answer_length"
    ]

    .mean()

)

summary_df = pd.DataFrame({

    "Metric": [

        "Total Queries",

        "Passed",

        "Needs Review",

        "Failed",

        "Average Top Similarity",

        "Average Answer Length"

    ],

    "Value": [

        total_queries,

        passed_queries,

        review_queries,

        failed_queries,

        round(
            average_top_similarity,
            4
        ),

        round(
            average_answer_length,
            2
        )

    ]

})

# ============================================================
# CREATE OUTPUT FOLDER
# ============================================================

os.makedirs(

    OUTPUT_FOLDER,

    exist_ok=True

)

# ============================================================
# SAVE EXCEL REPORT
# ============================================================

print(
    "\nSaving evaluation report..."
)

with pd.ExcelWriter(

    OUTPUT_FILE,

    engine="openpyxl"

) as writer:

    summary_df.to_excel(

        writer,

        sheet_name="Evaluation Summary",

        index=False

    )

    end_to_end_df.to_excel(

        writer,

        sheet_name="End-to-End Results",

        index=False

    )

    retrieval_metrics_df.to_excel(

        writer,

        sheet_name="Retrieval Metrics",

        index=False

    )

    llm_metrics_df.to_excel(

        writer,

        sheet_name="LLM Answer Metrics",

        index=False

    )

# ============================================================
# FORMAT EXCEL
# ============================================================

print(
    "Formatting Excel report..."
)

workbook = load_workbook(

    OUTPUT_FILE

)

for sheet in workbook.worksheets:

    sheet.freeze_panes = "A2"

    # Header formatting

    for cell in sheet[1]:

        cell.font = Font(

            bold=True

        )

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center"

        )

    # Wrap text

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(

                vertical="top",

                wrap_text=True

            )

    # Column widths

    for column_cells in sheet.columns:

        column_letter = get_column_letter(

            column_cells[0].column

        )

        max_length = 0

        for cell in column_cells:

            try:

                length = len(

                    str(
                        cell.value
                    )

                )

                if length > max_length:

                    max_length = length

            except:

                pass

        sheet.column_dimensions[
            column_letter
        ].width = min(

            max_length + 2,

            60

        )

workbook.save(

    OUTPUT_FILE

)

# ============================================================
# FINAL OUTPUT
# ============================================================

print("\n" + "=" * 70)

print(
    "RAG END-TO-END EVALUATION COMPLETED"
)

print("=" * 70)

print(
    f"\nTotal queries: "
    f"{total_queries}"
)

print(
    f"Passed: "
    f"{passed_queries}"
)

print(
    f"Needs review: "
    f"{review_queries}"
)

print(
    f"Failed: "
    f"{failed_queries}"
)

print(
    f"\nAverage top similarity: "
    f"{average_top_similarity:.4f}"
)

print(
    f"Average answer length: "
    f"{average_answer_length:.2f}"
)

print(
    "\nEvaluation report saved at:"
)

print(
    OUTPUT_FILE
)

print("\n" + "=" * 70)

print(
    "NEXT STEP"
)

print("=" * 70)

print(
    "\nReview the Excel report sheets:"
)

print(
    "1. Evaluation Summary"
)

print(
    "2. End-to-End Results"
)

print(
    "3. Retrieval Metrics"
)

print(
    "4. LLM Answer Metrics"
)

print(
    "\nAfter reviewing the report,"
)

print(
    "the next stage is RAG Optimization "
    "and Performance Improvement."
)