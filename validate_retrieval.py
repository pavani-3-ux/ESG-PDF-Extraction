import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = "output/retrieval/retrieval_similarity_search_results.xlsx"

OUTPUT_FOLDER = "output/validation"

OUTPUT_FILE = os.path.join(
    OUTPUT_FOLDER,
    "retrieval_validation_report.xlsx"
)


# ============================================================
# CREATE OUTPUT FOLDER
# ============================================================

os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# ============================================================
# START
# ============================================================

print("=" * 70)
print("RETRIEVAL VALIDATION")
print("=" * 70)


# ============================================================
# CHECK INPUT FILE
# ============================================================

if not os.path.exists(INPUT_FILE):

    print("\n❌ Retrieval result file not found!")

    print("\nExpected location:")
    print(INPUT_FILE)

    print("\nPlease make sure your retrieval process has created:")
    print("retrieval_similarity_search_results.xlsx")

    raise SystemExit


print("\n✅ Retrieval result file found!")
print("Location:", INPUT_FILE)


# ============================================================
# READ EXCEL FILE
# ============================================================

try:

    df = pd.read_excel(INPUT_FILE)

    print("\n✅ Retrieval results loaded successfully!")

    print("Rows:", len(df))
    print("Columns:", len(df.columns))

except Exception as e:

    print("\n❌ Error while reading Excel file!")

    print("Error:", e)

    raise SystemExit


# ============================================================
# DISPLAY COLUMNS
# ============================================================

print("\n" + "=" * 70)
print("INPUT COLUMNS")
print("=" * 70)

for column in df.columns:

    print("•", column)


# ============================================================
# NORMALIZE COLUMN NAMES
# ============================================================

df.columns = [
    str(column).strip().lower().replace(" ", "_")
    for column in df.columns
]


# ============================================================
# HELPER FUNCTION
# ============================================================

def find_column(possible_names):

    for name in possible_names:

        if name in df.columns:
            return name

    return None


# ============================================================
# DETECT IMPORTANT COLUMNS
# ============================================================

query_column = find_column([
    "query",
    "question",
    "user_query",
    "search_query"
])


rank_column = find_column([
    "rank",
    "result_rank"
])


distance_column = find_column([
    "distance",
    "similarity_distance"
])


company_column = find_column([
    "company",
    "source_company"
])


collection_column = find_column([
    "collection",
    "collection_name"
])


content_column = find_column([
    "content",
    "document",
    "text",
    "retrieved_content"
])


id_column = find_column([
    "id",
    "chunk_id",
    "document_id"
])


# ============================================================
# COLUMN DETECTION REPORT
# ============================================================

print("\n" + "=" * 70)
print("COLUMN DETECTION")
print("=" * 70)

print("Query column      :", query_column)
print("Rank column       :", rank_column)
print("Distance column   :", distance_column)
print("Company column    :", company_column)
print("Collection column :", collection_column)
print("Content column    :", content_column)
print("ID column         :", id_column)


# ============================================================
# CREATE VALIDATION DATAFRAME
# ============================================================

validation_df = df.copy()


# ============================================================
# BASIC CONTENT VALIDATION
# ============================================================

if content_column:

    validation_df["content_present"] = validation_df[
        content_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x) and str(x).strip() != ""
        else "NO"
    )

else:

    validation_df["content_present"] = "UNKNOWN"


# ============================================================
# COMPANY VALIDATION
# ============================================================

if company_column:

    validation_df["company_present"] = validation_df[
        company_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x) and str(x).strip() != ""
        else "NO"
    )

else:

    validation_df["company_present"] = "UNKNOWN"


# ============================================================
# COLLECTION VALIDATION
# ============================================================

if collection_column:

    validation_df["collection_present"] = validation_df[
        collection_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x) and str(x).strip() != ""
        else "NO"
    )

else:

    validation_df["collection_present"] = "UNKNOWN"


# ============================================================
# DISTANCE VALIDATION
# ============================================================

if distance_column:

    validation_df["distance_valid"] = validation_df[
        distance_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x)
        and isinstance(x, (int, float))
        else "NO"
    )

else:

    validation_df["distance_valid"] = "UNKNOWN"


# ============================================================
# RANK VALIDATION
# ============================================================

if rank_column:

    validation_df["rank_valid"] = validation_df[
        rank_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x)
        else "NO"
    )

else:

    validation_df["rank_valid"] = "UNKNOWN"


# ============================================================
# QUERY VALIDATION
# ============================================================

if query_column:

    validation_df["query_present"] = validation_df[
        query_column
    ].apply(
        lambda x:
        "YES"
        if pd.notna(x) and str(x).strip() != ""
        else "NO"
    )

else:

    validation_df["query_present"] = "UNKNOWN"


# ============================================================
# AUTOMATIC VALIDATION STATUS
# ============================================================

def calculate_status(row):

    checks = [
        row["query_present"],
        row["content_present"],
        row["company_present"],
        row["collection_present"],
        row["distance_valid"],
        row["rank_valid"]
    ]

    if "NO" in checks:

        return "FAILED"

    elif "UNKNOWN" in checks:

        return "REVIEW"

    else:

        return "PASSED"


validation_df["automatic_validation_status"] = validation_df.apply(
    calculate_status,
    axis=1
)


# ============================================================
# MANUAL VALIDATION COLUMNS
# ============================================================

validation_df["content_relevant"] = ""

validation_df["correct_company"] = ""

validation_df["correct_source"] = ""

validation_df["answers_query"] = ""

validation_df["retrieval_quality"] = ""

validation_df["review_comments"] = ""


# ============================================================
# REORDER COLUMNS
# ============================================================

preferred_columns = []

for column in [
    query_column,
    rank_column,
    distance_column,
    company_column,
    collection_column,
    id_column,
    content_column
]:

    if column and column in validation_df.columns:

        preferred_columns.append(column)


validation_columns = [

    "query_present",
    "content_present",
    "company_present",
    "collection_present",
    "distance_valid",
    "rank_valid",

    "automatic_validation_status",

    "content_relevant",
    "correct_company",
    "correct_source",
    "answers_query",
    "retrieval_quality",
    "review_comments"
]


remaining_columns = [
    column
    for column in validation_df.columns
    if column not in preferred_columns
    and column not in validation_columns
]


validation_df = validation_df[
    preferred_columns
    + remaining_columns
    + validation_columns
]


# ============================================================
# CREATE SUMMARY
# ============================================================

total_results = len(validation_df)

passed_count = (
    validation_df[
        "automatic_validation_status"
    ] == "PASSED"
).sum()


failed_count = (
    validation_df[
        "automatic_validation_status"
    ] == "FAILED"
).sum()


review_count = (
    validation_df[
        "automatic_validation_status"
    ] == "REVIEW"
).sum()


# ============================================================
# QUERY SUMMARY
# ============================================================

if query_column:

    query_summary = (
        validation_df
        .groupby(query_column)
        .agg(
            total_results=(
                query_column,
                "count"
            ),

            content_available=(
                "content_present",
                lambda x: (
                    x == "YES"
                ).sum()
            ),

            passed_results=(
                "automatic_validation_status",
                lambda x: (
                    x == "PASSED"
                ).sum()
            ),

            failed_results=(
                "automatic_validation_status",
                lambda x: (
                    x == "FAILED"
                ).sum()
            )
        )
        .reset_index()
    )

else:

    query_summary = pd.DataFrame()


# ============================================================
# SUMMARY DATA
# ============================================================

summary_df = pd.DataFrame({

    "Metric": [

        "Input File",

        "Total Retrieval Results",

        "Passed Results",

        "Failed Results",

        "Results Requiring Review",

        "Queries Found",

        "Content Present",

        "Content Missing"

    ],

    "Value": [

        INPUT_FILE,

        total_results,

        passed_count,

        failed_count,

        review_count,

        validation_df[
            query_column
        ].nunique()
        if query_column
        else "UNKNOWN",

        (
            validation_df[
                "content_present"
            ] == "YES"
        ).sum(),

        (
            validation_df[
                "content_present"
            ] == "NO"
        ).sum()

    ]

})


# ============================================================
# VALIDATION INSTRUCTIONS
# ============================================================

instructions_df = pd.DataFrame({

    "Validation Field": [

        "content_relevant",

        "correct_company",

        "correct_source",

        "answers_query",

        "retrieval_quality",

        "review_comments"

    ],

    "What To Enter": [

        "YES if retrieved content is relevant to the query, otherwise NO",

        "YES if the retrieved result belongs to the expected company, otherwise NO",

        "YES if the retrieved result comes from the correct document/source, otherwise NO",

        "YES if the retrieved content actually answers the user's query, otherwise NO",

        "GOOD / MEDIUM / POOR",

        "Write any comments about incorrect or irrelevant retrieval"

    ]

})


# ============================================================
# SAVE EXCEL FILE
# ============================================================

print("\n" + "=" * 70)
print("CREATING VALIDATION REPORT")
print("=" * 70)


with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    # Detailed validation
    validation_df.to_excel(
        writer,
        sheet_name="Retrieval Validation",
        index=False
    )

    # Overall summary
    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False
    )

    # Query-level summary
    if not query_summary.empty:

        query_summary.to_excel(
            writer,
            sheet_name="Query Summary",
            index=False
        )

    # Instructions
    instructions_df.to_excel(
        writer,
        sheet_name="Validation Instructions",
        index=False
    )


# ============================================================
# FORMAT EXCEL FILE
# ============================================================

try:

    workbook = load_workbook(
        OUTPUT_FILE
    )


    # --------------------------------------------------------
    # STYLES
    # --------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )


    # --------------------------------------------------------
    # FORMAT ALL SHEETS
    # --------------------------------------------------------

    for worksheet in workbook.worksheets:

        # Freeze first row
        worksheet.freeze_panes = "A2"


        # Header formatting
        for cell in worksheet[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            cell.border = thin_border


        # Format all cells
        for row in worksheet.iter_rows():

            for cell in row:

                cell.border = thin_border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )


        # Adjust column widths
        for column_cells in worksheet.columns:

            column_letter = get_column_letter(
                column_cells[0].column
            )

            max_length = 0

            for cell in column_cells:

                try:

                    cell_length = len(
                        str(cell.value)
                    )

                    if cell_length > max_length:

                        max_length = cell_length

                except:

                    pass


            # Limit very large columns
            adjusted_width = min(
                max(max_length + 2, 12),
                60
            )

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width


        # Increase header height
        worksheet.row_dimensions[1].height = 30


    # --------------------------------------------------------
    # SAVE FORMATTED WORKBOOK
    # --------------------------------------------------------

    workbook.save(
        OUTPUT_FILE
    )


    print("\n✅ Excel formatting completed!")


except Exception as e:

    print("\n⚠️ Excel formatting warning:")
    print(e)


# ============================================================
# FINAL OUTPUT
# ============================================================

print("\n" + "=" * 70)
print("RETRIEVAL VALIDATION COMPLETED")
print("=" * 70)

print("\nTotal retrieval results:", total_results)

print("Passed:", passed_count)

print("Failed:", failed_count)

print("Needs review:", review_count)

print("\nValidation report saved at:")

print(OUTPUT_FILE)

print("\n" + "=" * 70)
print("NEXT STEP")
print("=" * 70)

print(
    "Open the Excel report and manually review:"
)

print(
    "1. content_relevant"
)

print(
    "2. correct_company"
)

print(
    "3. correct_source"
)

print(
    "4. answers_query"
)

print(
    "5. retrieval_quality"
)

print(
    "6. review_comments"
)

print("\nAfter completing the manual validation,")
print("the next stage is Retrieval Quality Analysis / Reranking.")