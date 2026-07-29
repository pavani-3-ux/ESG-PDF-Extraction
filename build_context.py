import os
import pandas as pd
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = (
    "output/reranking/"
    "reranked_results.xlsx"
)

OUTPUT_DIR = "output/context"

OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    "built_context.xlsx"
)

# Number of top reranked results to keep for each query
TOP_K = 5


# ============================================================
# HELPER FUNCTION
# ============================================================

def find_column(df, possible_names):

    """
    Find a column using multiple possible column names.
    """

    normalized_columns = {
        str(column).strip().lower(): column
        for column in df.columns
    }

    for name in possible_names:

        name_lower = (
            name.strip().lower()
        )

        if name_lower in normalized_columns:

            return normalized_columns[
                name_lower
            ]

    return None


# ============================================================
# START
# ============================================================

print("=" * 70)
print("CONTEXT BUILDING")
print("=" * 70)


# ============================================================
# CHECK INPUT FILE
# ============================================================

if not os.path.exists(INPUT_FILE):

    print()
    print(
        "❌ Reranked results file not found!"
    )

    print()
    print(
        "Expected location:"
    )

    print(
        INPUT_FILE
    )

    print()
    print(
        "Please run rerank_results.py first."
    )

    raise SystemExit(1)


print()
print(
    "✅ Reranked results file found!"
)

print(
    "Location:",
    INPUT_FILE
)


# ============================================================
# LOAD RERANKED RESULTS
# ============================================================

try:

    df = pd.read_excel(
        INPUT_FILE
    )

    print()
    print(
        "✅ Reranked results loaded successfully!"
    )

    print(
        "Total rows:",
        len(df)
    )

    print(
        "Total columns:",
        len(df.columns)
    )

except Exception as e:

    print()
    print(
        "❌ Error loading reranked results!"
    )

    print(
        "Error:",
        e
    )

    raise SystemExit(1)


# ============================================================
# DISPLAY INPUT COLUMNS
# ============================================================

print()
print("=" * 70)
print("INPUT COLUMNS")
print("=" * 70)

for column in df.columns:

    print(
        "•",
        column
    )


# ============================================================
# DETECT COLUMNS
# ============================================================

query_col = find_column(
    df,
    [
        "query",
        "question",
        "user_query"
    ]
)


content_col = find_column(
    df,
    [
        "content",
        "text",
        "document",
        "chunk_content"
    ]
)


reranked_rank_col = find_column(
    df,
    [
        "reranked_rank",
        "new_rank"
    ]
)


reranker_score_col = find_column(
    df,
    [
        "reranker_score",
        "rerank_score",
        "reranking_score"
    ]
)


content_type_col = find_column(
    df,
    [
        "content_type",
        "type"
    ]
)


company_col = find_column(
    df,
    [
        "company",
        "company_name"
    ]
)


document_col = find_column(
    df,
    [
        "document_name",
        "document"
    ]
)


source_col = find_column(
    df,
    [
        "source_file",
        "source"
    ]
)


section_col = find_column(
    df,
    [
        "section",
        "section_name"
    ]
)


chunk_id_col = find_column(
    df,
    [
        "chunk_id",
        "result_id",
        "id"
    ]
)


chunk_index_col = find_column(
    df,
    [
        "chunk_index",
        "index"
    ]
)


# ============================================================
# PRINT DETECTED COLUMNS
# ============================================================

print()
print("=" * 70)
print("COLUMN DETECTION")
print("=" * 70)

print(
    "Query column          :",
    query_col
)

print(
    "Content column        :",
    content_col
)

print(
    "Reranked rank         :",
    reranked_rank_col
)

print(
    "Reranker score        :",
    reranker_score_col
)

print(
    "Content type          :",
    content_type_col
)

print(
    "Company column        :",
    company_col
)

print(
    "Document column       :",
    document_col
)

print(
    "Source column         :",
    source_col
)

print(
    "Section column        :",
    section_col
)

print(
    "Chunk ID column       :",
    chunk_id_col
)


# ============================================================
# VALIDATE REQUIRED COLUMNS
# ============================================================

if query_col is None:

    print()
    print(
        "❌ Query column not found!"
    )

    raise SystemExit(1)


if content_col is None:

    print()
    print(
        "❌ Content column not found!"
    )

    raise SystemExit(1)


if reranked_rank_col is None:

    print()
    print(
        "❌ Reranked rank column not found!"
    )

    raise SystemExit(1)


# ============================================================
# CLEAN DATA
# ============================================================

print()
print("=" * 70)
print("CLEANING CONTEXT DATA")
print("=" * 70)


context_df = df.copy()


# Remove rows without query

context_df = context_df[
    context_df[
        query_col
    ].notna()
]


# Remove rows without content

context_df = context_df[
    context_df[
        content_col
    ].notna()
]


# Convert content to string

context_df[
    content_col
] = (

    context_df[
        content_col
    ]

    .astype(str)

    .str.strip()

)


# Remove empty content

context_df = context_df[

    context_df[
        content_col
    ] != ""

]


print()
print(
    "Rows after cleaning:",
    len(context_df)
)


# ============================================================
# CONVERT RERANKED RANK TO NUMERIC
# ============================================================

context_df[
    reranked_rank_col
] = pd.to_numeric(

    context_df[
        reranked_rank_col
    ],

    errors="coerce"

)


# Remove rows with invalid rank

context_df = context_df[

    context_df[
        reranked_rank_col
    ].notna()

]


# ============================================================
# SORT BY QUERY AND RERANKED RANK
# ============================================================

context_df = (

    context_df

    .sort_values(

        by=[
            query_col,
            reranked_rank_col
        ],

        ascending=[
            True,
            True
        ]

    )

)


# ============================================================
# SELECT TOP-K RESULTS
# ============================================================

print()
print("=" * 70)
print("SELECTING TOP-K RESULTS")
print("=" * 70)

print()
print(
    "Top-K:",
    TOP_K
)


top_k_df = (

    context_df

    .groupby(

        query_col,

        group_keys=False

    )

    .head(
        TOP_K
    )

    .copy()

)


print()
print(
    "Total context records:",
    len(top_k_df)
)


# ============================================================
# CREATE CONTEXT RANK
# ============================================================

top_k_df[
    "context_rank"
] = (

    top_k_df

    .groupby(
        query_col
    )

    .cumcount()

    + 1

)


# ============================================================
# IDENTIFY CONTENT TYPE
# ============================================================

if content_type_col is not None:

    top_k_df[
        "normalized_content_type"
    ] = (

        top_k_df[
            content_type_col
        ]

        .astype(str)

        .str.lower()

        .str.strip()

    )

else:

    top_k_df[
        "normalized_content_type"
    ] = "unknown"


# ============================================================
# CREATE CONTEXT TYPE LABEL
# ============================================================

def classify_content_type(value):

    value = str(
        value
    ).lower().strip()


    if "text" in value:

        return "TEXT"


    if "table" in value:

        return "TABLE"


    if "image" in value:

        return "IMAGE"


    return "OTHER"


top_k_df[
    "context_type"
] = (

    top_k_df[
        "normalized_content_type"
    ]

    .apply(
        classify_content_type
    )

)


# ============================================================
# CREATE STRUCTURED CONTEXT
# ============================================================

print()
print("=" * 70)
print("BUILDING STRUCTURED CONTEXT")
print("=" * 70)


def create_context_text(row):

    query = str(

        row.get(
            query_col,
            ""
        )

    )


    content = str(

        row.get(
            content_col,
            ""
        )

    )


    context_type = str(

        row.get(
            "context_type",
            "OTHER"
        )

    )


    rank = row.get(

        "context_rank",

        ""

    )


    score = row.get(

        reranker_score_col,

        ""

    )


    company = ""


    if company_col is not None:

        company = str(

            row.get(
                company_col,
                ""
            )

        )


    document = ""


    if document_col is not None:

        document = str(

            row.get(
                document_col,
                ""
            )

        )


    section = ""


    if section_col is not None:

        section = str(

            row.get(
                section_col,
                ""
            )

        )


    context_block = (

        f"[CONTEXT ITEM {rank}]\n"

        f"Content Type: {context_type}\n"

        f"Company: {company}\n"

        f"Document: {document}\n"

        f"Section: {section}\n"

        f"Reranker Score: {score}\n"

        f"Content:\n"

        f"{content}\n"

        f"[END CONTEXT ITEM]"

    )


    return context_block


top_k_df[
    "structured_context"
] = (

    top_k_df

    .apply(

        create_context_text,

        axis=1

    )

)


# ============================================================
# BUILD COMBINED CONTEXT PER QUERY
# ============================================================

print()
print(
    "Combining context items for each query..."
)


combined_context_rows = []


for query, group in top_k_df.groupby(
    query_col
):

    context_items = (

        group[
            "structured_context"
        ]

        .tolist()

    )


    combined_context = (

        "\n\n"

        .join(
            context_items
        )

    )


    # --------------------------------------------------------
    # COUNT CONTENT TYPES
    # --------------------------------------------------------

    text_count = (

        group[
            "context_type"
        ]

        .eq(
            "TEXT"
        )

        .sum()

    )


    table_count = (

        group[
            "context_type"
        ]

        .eq(
            "TABLE"
        )

        .sum()

    )


    image_count = (

        group[
            "context_type"
        ]

        .eq(
            "IMAGE"
        )

        .sum()

    )


    combined_context_rows.append(

        {

            "query": query,

            "total_context_items": len(
                group
            ),

            "text_items": text_count,

            "table_items": table_count,

            "image_items": image_count,

            "combined_context": combined_context

        }

    )


combined_context_df = pd.DataFrame(

    combined_context_rows

)


# ============================================================
# CREATE CONTEXT METADATA
# ============================================================

context_metadata_columns = []


if query_col is not None:

    context_metadata_columns.append(
        query_col
    )


if company_col is not None:

    context_metadata_columns.append(
        company_col
    )


if document_col is not None:

    context_metadata_columns.append(
        document_col
    )


if source_col is not None:

    context_metadata_columns.append(
        source_col
    )


if section_col is not None:

    context_metadata_columns.append(
        section_col
    )


if chunk_id_col is not None:

    context_metadata_columns.append(
        chunk_id_col
    )


if content_type_col is not None:

    context_metadata_columns.append(
        content_type_col
    )


if reranked_rank_col is not None:

    context_metadata_columns.append(
        reranked_rank_col
    )


if reranker_score_col is not None:

    context_metadata_columns.append(
        reranker_score_col
    )


context_metadata_columns.extend(

    [

        "context_rank",

        "context_type",

        content_col,

        "structured_context"

    ]

)


# Keep only columns that exist

context_metadata_columns = [

    column

    for column in context_metadata_columns

    if column in top_k_df.columns

]


context_metadata_df = top_k_df[

    context_metadata_columns

].copy()


# ============================================================
# CREATE OUTPUT DIRECTORY
# ============================================================

os.makedirs(

    OUTPUT_DIR,

    exist_ok=True

)


# ============================================================
# SAVE CONTEXT TO EXCEL
# ============================================================

print()
print("=" * 70)
print("SAVING BUILT CONTEXT")
print("=" * 70)


try:

    with pd.ExcelWriter(

        OUTPUT_FILE,

        engine="openpyxl"

    ) as writer:


        # ----------------------------------------------------
        # SHEET 1: COMBINED CONTEXT
        # ----------------------------------------------------

        combined_context_df.to_excel(

            writer,

            sheet_name="Combined Context",

            index=False

        )


        # ----------------------------------------------------
        # SHEET 2: CONTEXT ITEMS
        # ----------------------------------------------------

        context_metadata_df.to_excel(

            writer,

            sheet_name="Context Items",

            index=False

        )


        # ----------------------------------------------------
        # SHEET 3: CONTEXT SUMMARY
        # ----------------------------------------------------

        summary_data = [

            [

                "Total Queries",

                len(
                    combined_context_df
                )

            ],

            [

                "Total Context Items",

                len(
                    top_k_df
                )

            ],

            [

                "Top K Per Query",

                TOP_K

            ],

            [

                "Total Text Items",

                (
                    top_k_df[
                        "context_type"
                    ]

                    .eq(
                        "TEXT"
                    )

                    .sum()
                )

            ],

            [

                "Total Table Items",

                (
                    top_k_df[
                        "context_type"
                    ]

                    .eq(
                        "TABLE"
                    )

                    .sum()
                )

            ],

            [

                "Total Image Items",

                (
                    top_k_df[
                        "context_type"
                    ]

                    .eq(
                        "IMAGE"
                    )

                    .sum()
                )

            ]

        ]


        summary_df = pd.DataFrame(

            summary_data,

            columns=[

                "Metric",

                "Value"

            ]

        )


        summary_df.to_excel(

            writer,

            sheet_name="Context Summary",

            index=False

        )


    print()
    print(
        "✅ Context Excel file created successfully!"
    )


except Exception as e:

    print()
    print(
        "❌ Error creating context file!"
    )

    print(
        "Error:",
        e
    )

    raise SystemExit(1)


# ============================================================
# FORMAT EXCEL
# ============================================================

print()
print(
    "Formatting Excel file..."
)


try:

    workbook = load_workbook(

        OUTPUT_FILE

    )


    for worksheet in workbook.worksheets:


        # Freeze top row

        worksheet.freeze_panes = "A2"


        # Enable text wrapping

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = (

                    cell.alignment.copy(

                        wrap_text=True

                    )

                )


        # Auto-adjust column widths

        for column_cells in worksheet.columns:

            max_length = 0


            column_letter = (

                column_cells[
                    0
                ]

                .column_letter

            )


            for cell in column_cells:

                try:

                    cell_length = len(

                        str(
                            cell.value
                        )

                    )


                    if cell_length > max_length:

                        max_length = cell_length


                except:

                    pass


            worksheet.column_dimensions[
                column_letter
            ].width = min(

                max_length + 2,

                60

            )


    workbook.save(

        OUTPUT_FILE

    )


    print(
        "✅ Excel formatting completed!"
    )


except Exception as e:

    print(
        "⚠️ Excel formatting warning:",
        e
    )


# ============================================================
# FINAL SUMMARY
# ============================================================

print()
print("=" * 70)
print("CONTEXT BUILDING COMPLETED SUCCESSFULLY")
print("=" * 70)

print()

print(
    "Total queries:",
    len(
        combined_context_df
    )
)

print(
    "Total context items:",
    len(
        top_k_df
    )
)

print(
    "Top-K per query:",
    TOP_K
)

print()

print(
    "Output file:"
)

print(
    OUTPUT_FILE
)

print()
print("=" * 70)
print("NEXT STEP")
print("=" * 70)

print()

print(
    "Next step:"
)

print(
    "Create validate_context.py"
)

print()

print(
    "This will validate the built context before"
)

print(
    "sending it to the LLM / RAG answer generation stage."
)

print("=" * 70)