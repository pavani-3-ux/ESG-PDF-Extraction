import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# 1. PROJECT PATH
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

# Chunked data folders
TEXT_CHUNKS_DIR = BASE_DIR / "output" / "chunks" / "text"
TABLE_CHUNKS_DIR = BASE_DIR / "output" / "chunks" / "tables"
IMAGE_CHUNKS_DIR = BASE_DIR / "output" / "chunks" / "images"

# Validation output folder
VALIDATION_DIR = BASE_DIR / "output" / "validation"

# Create validation folder if it does not exist
VALIDATION_DIR.mkdir(parents=True, exist_ok=True)

# Final Excel file
OUTPUT_FILE = VALIDATION_DIR / "chunking_validation_report.xlsx"


# ============================================================
# 2. VALIDATE TEXT CHUNKS
# ============================================================

def validate_text_chunks():

    results = []

    if not TEXT_CHUNKS_DIR.exists():
        results.append([
            "Text",
            "Folder Check",
            "FAIL",
            "Text chunks folder does not exist",
            0,
            0
        ])
        return results

    files = list(TEXT_CHUNKS_DIR.rglob("*.txt"))

    if not files:
        results.append([
            "Text",
            "File Check",
            "FAIL",
            "No text chunk files found",
            0,
            0
        ])
        return results

    for file in files:

        try:
            content = file.read_text(
                encoding="utf-8",
                errors="ignore"
            )

            content = content.strip()

            if len(content) == 0:

                results.append([
                    "Text",
                    file.name,
                    "FAIL",
                    "Empty chunk",
                    0,
                    0
                ])

            else:

                results.append([
                    "Text",
                    file.name,
                    "PASS",
                    "Valid text chunk",
                    len(content),
                    len(content.split())
                ])

        except Exception as e:

            results.append([
                "Text",
                file.name,
                "FAIL",
                f"Error reading file: {e}",
                0,
                0
            ])

    return results


# ============================================================
# 3. VALIDATE TABLE CHUNKS
# ============================================================

def validate_table_chunks():

    results = []

    if not TABLE_CHUNKS_DIR.exists():

        results.append([
            "Table",
            "Folder Check",
            "FAIL",
            "Table chunks folder does not exist",
            0,
            0
        ])

        return results

    files = list(TABLE_CHUNKS_DIR.rglob("*.xlsx"))

    if not files:

        results.append([
            "Table",
            "File Check",
            "FAIL",
            "No table chunk Excel files found",
            0,
            0
        ])

        return results

    for file in files:

        try:

            from openpyxl import load_workbook

            workbook = load_workbook(
                file,
                read_only=True,
                data_only=True
            )

            total_rows = 0
            total_columns = 0

            for sheet in workbook.worksheets:

                rows = list(sheet.iter_rows(values_only=True))

                if rows:

                    total_rows += len(rows)

                    max_columns = max(
                        len(row) for row in rows
                    )

                    total_columns = max(
                        total_columns,
                        max_columns
                    )

            workbook.close()

            if total_rows == 0:

                results.append([
                    "Table",
                    file.name,
                    "FAIL",
                    "Empty table chunk",
                    0,
                    0
                ])

            else:

                results.append([
                    "Table",
                    file.name,
                    "PASS",
                    "Valid table chunk",
                    total_rows,
                    total_columns
                ])

        except Exception as e:

            results.append([
                "Table",
                file.name,
                "FAIL",
                f"Error reading Excel file: {e}",
                0,
                0
            ])

    return results


# ============================================================
# 4. VALIDATE IMAGE CHUNKS
# ============================================================

def validate_image_chunks():

    results = []

    if not IMAGE_CHUNKS_DIR.exists():

        results.append([
            "Image",
            "Folder Check",
            "FAIL",
            "Image chunks folder does not exist",
            0,
            0
        ])

        return results

    image_extensions = [
        "*.png",
        "*.jpg",
        "*.jpeg",
        "*.webp"
    ]

    files = []

    for extension in image_extensions:

        files.extend(
            IMAGE_CHUNKS_DIR.rglob(extension)
        )

    if not files:

        results.append([
            "Image",
            "File Check",
            "FAIL",
            "No image chunk files found",
            0,
            0
        ])

        return results

    for file in files:

        try:

            file_size = file.stat().st_size

            if file_size == 0:

                results.append([
                    "Image",
                    file.name,
                    "FAIL",
                    "Empty image file",
                    0,
                    0
                ])

            else:

                results.append([
                    "Image",
                    file.name,
                    "PASS",
                    "Valid image chunk",
                    file_size,
                    0
                ])

        except Exception as e:

            results.append([
                "Image",
                file.name,
                "FAIL",
                f"Error checking image: {e}",
                0,
                0
            ])

    return results


# ============================================================
# 5. CREATE EXCEL VALIDATION REPORT
# ============================================================

def create_excel_report(all_results):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Chunking Validation"

    # Header
    headers = [
        "Data Type",
        "File Name",
        "Status",
        "Validation Result",
        "Size / Rows",
        "Words / Columns"
    ]

    worksheet.append(headers)

    # Header formatting
    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    # Add validation results
    for result in all_results:

        worksheet.append(result)

    # Format status cells
    for row in worksheet.iter_rows(
        min_row=2,
        max_col=6
    ):

        status_cell = row[2]

        if status_cell.value == "PASS":

            status_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="C6EFCE"
            )

        elif status_cell.value == "FAIL":

            status_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FFC7CE"
            )

    # Adjust column widths
    column_widths = {
        "A": 15,
        "B": 45,
        "C": 12,
        "D": 45,
        "E": 18,
        "F": 18
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # Freeze header
    worksheet.freeze_panes = "A2"

    # Save Excel report
    workbook.save(OUTPUT_FILE)

    print("\n========================================")
    print("CHUNKING VALIDATION COMPLETED")
    print("========================================")

    print(f"\nExcel Report:")
    print(OUTPUT_FILE)


# ============================================================
# 6. MAIN PROGRAM
# ============================================================

if __name__ == "__main__":

    print("\nStarting Chunking Validation...")

    print("\nValidating Text Chunks...")
    text_results = validate_text_chunks()

    print("Validating Table Chunks...")
    table_results = validate_table_chunks()

    print("Validating Image Chunks...")
    image_results = validate_image_chunks()

    # Combine all results
    all_results = (
        text_results
        + table_results
        + image_results
    )

    # Create Excel report
    create_excel_report(all_results)