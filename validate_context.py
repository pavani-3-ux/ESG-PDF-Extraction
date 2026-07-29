import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = (
    "output/context/"
    "built_context.xlsx"
)

OUTPUT_DIR = "output/validation"

OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    "context_validation_report.xlsx"
)

# Expected maximum number of context items per query
TOP_K = 5


# ============================================================
# HELPER FUNCTION
# ============================================================

def find_column(df, possible_names):

    """
    Find a column using multiple possible column names.
    """

    normalized_columns = {
        str(column).strip().lower(): column
        for column in df.columns
    }

    for name in possible_names:

        name_lower = (
            name.strip().lower()
        )

        if name_lower in normalized_columns:

            return normalized_columns[
                name_lower
            ]

    return None


# ============================================================
# START
# ============================================================

print("=" * 70)
print("CONTEXT VALIDATION")
print("=" * 70)


# ============================================================
# CHECK INPUT FILE
# ============================================================

if not os.path.exists(INPUT_FILE):

    print()
    print(
        "❌ Context file not found!"
    )

    print()
    print(
        "Expected location:"
    )

    print(
        INPUT_FILE
    )

    print()
    print(
        "Please run build_context.py first."
    )

    raise SystemExit(1)


print()
print(
    "✅ Context file found!"
)

print(
    "Location:",
    INPUT_FILE
)


# ============================================================
# LOAD EXCEL WORKBOOK
# ============================================================

try:

    excel_file = pd.ExcelFile(
        INPUT_FILE
    )

    available_sheets = (
        excel_file.sheet_names
    )

    print()
    print(
        "✅ Excel workbook loaded successfully!"
    )

    print()
    print(
        "Available sheets:"
    )

    for sheet in available_sheets:

        print(
            "•",
            sheet
        )

except Exception as e:

    print()
    print(
        "❌ Error loading Excel workbook!"
    )

    print(
        "Error:",
        e
    )

    raise SystemExit(1)


# ============================================================
# EXPECTED SHEETS
# ============================================================

expected_sheets = [

    "Combined Context",

    "Context Items",

    "Context Summary"

]


# ============================================================
# SHEET VALIDATION
# ============================================================

print()
print("=" * 70)
print("SHEET VALIDATION")
print("=" * 70)


sheet_validation = []


for sheet in expected_sheets:

    if sheet in available_sheets:

        status = "PASS"

        print(
            f"✅ Found sheet: {sheet}"
        )

    else:

        status = "FAIL"

        print(
            f"❌ Missing sheet: {sheet}"
        )


    sheet_validation.append(

        [

            sheet,

            status

        ]

    )


sheet_validation_df = pd.DataFrame(

    sheet_validation,

    columns=[

        "Sheet Name",

        "Status"

    ]

)


# ============================================================
# LOAD COMBINED CONTEXT
# ============================================================

if "Combined Context" in available_sheets:

    combined_df = pd.read_excel(

        INPUT_FILE,

        sheet_name="Combined Context"

    )

else:

    combined_df = pd.DataFrame()


# ============================================================
# LOAD CONTEXT ITEMS
# ============================================================

if "Context Items" in available_sheets:

    items_df = pd.read_excel(

        INPUT_FILE,

        sheet_name="Context Items"

    )

else:

    items_df = pd.DataFrame()


# ============================================================
# LOAD CONTEXT SUMMARY
# ============================================================

if "Context Summary" in available_sheets:

    summary_df = pd.read_excel(

        INPUT_FILE,

        sheet_name="Context Summary"

    )

else:

    summary_df = pd.DataFrame()


# ============================================================
# DISPLAY DATA INFORMATION
# ============================================================

print()
print("=" * 70)
print("CONTEXT DATA INFORMATION")
print("=" * 70)

print()

print(
    "Combined Context rows:",
    len(combined_df)
)

print(
    "Context Items rows:",
    len(items_df)
)

print(
    "Context Summary rows:",
    len(summary_df)
)


# ============================================================
# DETECT COLUMNS - COMBINED CONTEXT
# ============================================================

combined_query_col = find_column(

    combined_df,

    [

        "query",

        "question",

        "user_query"

    ]

)


combined_context_col = find_column(

    combined_df,

    [

        "combined_context",

        "context",

        "built_context"

    ]

)


total_items_col = find_column(

    combined_df,

    [

        "total_context_items",

        "context_items",

        "total_items"

    ]

)


text_items_col = find_column(

    combined_df,

    [

        "text_items",

        "text_count"

    ]

)


table_items_col = find_column(

    combined_df,

    [

        "table_items",

        "table_count"

    ]

)


image_items_col = find_column(

    combined_df,

    [

        "image_items",

        "image_count"

    ]

)


# ============================================================
# DETECT COLUMNS - CONTEXT ITEMS
# ============================================================

items_query_col = find_column(

    items_df,

    [

        "query",

        "question",

        "user_query"

    ]

)


items_content_col = find_column(

    items_df,

    [

        "content",

        "text",

        "chunk_content"

    ]

)


items_type_col = find_column(

    items_df,

    [

        "context_type",

        "content_type",

        "type"

    ]

)


items_rank_col = find_column(

    items_df,

    [

        "context_rank",

        "reranked_rank",

        "rank"

    ]

)


items_score_col = find_column(

    items_df,

    [

        "reranker_score",

        "rerank_score",

        "reranking_score"

    ]

)


items_chunk_id_col = find_column(

    items_df,

    [

        "chunk_id",

        "result_id",

        "id"

    ]

)


# ============================================================
# PRINT COLUMN DETECTION
# ============================================================

print()
print("=" * 70)
print("COLUMN DETECTION")
print("=" * 70)

print()
print("COMBINED CONTEXT")

print(
    "Query column          :",
    combined_query_col
)

print(
    "Combined context      :",
    combined_context_col
)

print(
    "Total items column    :",
    total_items_col
)

print(
    "Text items column     :",
    text_items_col
)

print(
    "Table items column    :",
    table_items_col
)

print(
    "Image items column    :",
    image_items_col
)


print()
print("CONTEXT ITEMS")

print(
    "Query column          :",
    items_query_col
)

print(
    "Content column        :",
    items_content_col
)

print(
    "Content type column   :",
    items_type_col
)

print(
    "Context rank column   :",
    items_rank_col
)

print(
    "Reranker score column :",
    items_score_col
)

print(
    "Chunk ID column       :",
    items_chunk_id_col
)


# ============================================================
# CREATE VALIDATION DATAFRAME
# ============================================================

if not items_df.empty:

    validation_df = items_df.copy()

else:

    validation_df = pd.DataFrame()


# ============================================================
# VALIDATION 1
# QUERY PRESENT
# ============================================================

if (
    not validation_df.empty
    and
    items_query_col is not None
):

    validation_df[
        "query_present"
    ] = (

        validation_df[
            items_query_col
        ]

        .notna()

        &

        (

            validation_df[
                items_query_col
            ]

            .astype(str)

            .str.strip()

            != ""

        )

    )

else:

    validation_df[
        "query_present"
    ] = False


# ============================================================
# VALIDATION 2
# CONTENT PRESENT
# ============================================================

if (
    not validation_df.empty
    and
    items_content_col is not None
):

    validation_df[
        "content_present"
    ] = (

        validation_df[
            items_content_col
        ]

        .notna()

        &

        (

            validation_df[
                items_content_col
            ]

            .astype(str)

            .str.strip()

            != ""

        )

    )

else:

    validation_df[
        "content_present"
    ] = False


# ============================================================
# VALIDATION 3
# CONTENT TYPE PRESENT
# ============================================================

if (
    not validation_df.empty
    and
    items_type_col is not None
):

    validation_df[
        "content_type_present"
    ] = (

        validation_df[
            items_type_col
        ]

        .notna()

        &

        (

            validation_df[
                items_type_col
            ]

            .astype(str)

            .str.strip()

            != ""

        )

    )

else:

    validation_df[
        "content_type_present"
    ] = False


# ============================================================
# VALIDATION 4
# CONTEXT RANK PRESENT
# ============================================================

if (
    not validation_df.empty
    and
    items_rank_col is not None
):

    validation_df[
        "rank_present"
    ] = (

        validation_df[
            items_rank_col
        ]

        .notna()

    )

else:

    validation_df[
        "rank_present"
    ] = False


# ============================================================
# VALIDATION 5
# RERANKER SCORE PRESENT
# ============================================================

if (
    not validation_df.empty
    and
    items_score_col is not None
):

    validation_df[
        "reranker_score_present"
    ] = (

        validation_df[
            items_score_col
        ]

        .notna()

    )

else:

    validation_df[
        "reranker_score_present"
    ] = False


# ============================================================
# VALIDATION 6
# CHECK EMPTY CONTEXT
# ============================================================

if (
    not combined_df.empty
    and
    combined_context_col is not None
):

    combined_df[
        "context_present"
    ] = (

        combined_df[
            combined_context_col
        ]

        .notna()

        &

        (

            combined_df[
                combined_context_col
            ]

            .astype(str)

            .str.strip()

            != ""

        )

    )

else:

    combined_df[
        "context_present"
    ] = False


# ============================================================
# VALIDATION 7
# CHECK CONTEXT LENGTH
# ============================================================

if (
    not combined_df.empty
    and
    combined_context_col is not None
):

    combined_df[
        "context_length"
    ] = (

        combined_df[
            combined_context_col
        ]

        .astype(str)

        .str.len()

    )

    combined_df[
        "context_length_valid"
    ] = (

        combined_df[
            "context_length"
        ]

        > 0

    )

else:

    combined_df[
        "context_length"
    ] = 0

    combined_df[
        "context_length_valid"
    ] = False


# ============================================================
# VALIDATION 8
# CHECK TOP-K LIMIT
# ============================================================

if (
    not combined_df.empty
    and
    total_items_col is not None
):

    combined_df[
        "top_k_valid"
    ] = (

        pd.to_numeric(

            combined_df[
                total_items_col
            ],

            errors="coerce"

        )

        <= TOP_K

    )

else:

    combined_df[
        "top_k_valid"
    ] = False


# ============================================================
# VALIDATION 9
# CHECK DUPLICATE CHUNK IDs
# ============================================================

if (
    not validation_df.empty
    and
    items_chunk_id_col is not None
):

    validation_df[
        "duplicate_chunk_id"
    ] = (

        validation_df[
            items_chunk_id_col
        ]

        .duplicated(

            keep=False

        )

    )

else:

    validation_df[
        "duplicate_chunk_id"
    ] = False


# ============================================================
# VALIDATION 10
# CHECK CONTENT TYPES
# ============================================================

valid_content_types = [

    "TEXT",

    "TABLE",

    "IMAGE",

    "OTHER"

]


if (
    not validation_df.empty
    and
    items_type_col is not None
):

    validation_df[
        "content_type_valid"
    ] = (

        validation_df[
            items_type_col
        ]

        .astype(str)

        .str.upper()

        .isin(

            valid_content_types

        )

    )

else:

    validation_df[
        "content_type_valid"
    ] = False


# ============================================================
# VALIDATION 11
# CHECK CONTEXT RANK SEQUENCE
# ============================================================

validation_df[
    "rank_sequence_valid"
] = True


if (
    not validation_df.empty
    and
    items_query_col is not None
    and
    items_rank_col is not None
):

    for query in (

        validation_df[
            items_query_col
        ]

        .dropna()

        .unique()

    ):

        query_mask = (

            validation_df[
                items_query_col
            ]

            == query

        )


        query_ranks = (

            pd.to_numeric(

                validation_df.loc[

                    query_mask,

                    items_rank_col

                ],

                errors="coerce"

            )

            .dropna()

            .sort_values()

            .tolist()

        )


        expected_ranks = list(

            range(

                1,

                len(
                    query_ranks
                )

                + 1

            )

        )


        if query_ranks != expected_ranks:

            validation_df.loc[

                query_mask,

                "rank_sequence_valid"

            ] = False


# ============================================================
# OVERALL ITEM VALIDATION
# ============================================================

if not validation_df.empty:

    validation_df[
        "automatic_validation"
    ] = "PASS"


    validation_df.loc[

        (

            ~validation_df[
                "query_present"
            ]

            |

            ~validation_df[
                "content_present"
            ]

            |

            ~validation_df[
                "content_type_present"
            ]

            |

            ~validation_df[
                "rank_present"
            ]

            |

            ~validation_df[
                "reranker_score_present"
            ]

            |

            validation_df[
                "duplicate_chunk_id"
            ]

            |

            ~validation_df[
                "content_type_valid"
            ]

            |

            ~validation_df[
                "rank_sequence_valid"
            ]

        ),

        "automatic_validation"

    ] = "FAIL"

else:

    validation_df[
        "automatic_validation"
    ] = "FAIL"


# ============================================================
# COMBINED CONTEXT VALIDATION
# ============================================================

if not combined_df.empty:

    combined_df[
        "automatic_validation"
    ] = "PASS"


    combined_df.loc[

        (

            ~combined_df[
                "context_present"
            ]

            |

            ~combined_df[
                "context_length_valid"
            ]

            |

            ~combined_df[
                "top_k_valid"
            ]

        ),

        "automatic_validation"

    ] = "FAIL"

else:

    combined_df[
        "automatic_validation"
    ] = "FAIL"


# ============================================================
# SUMMARY CALCULATIONS
# ============================================================

total_context_items = len(
    validation_df
)


passed_items = (

    validation_df[
        "automatic_validation"
    ]

    == "PASS"

).sum()


failed_items = (

    validation_df[
        "automatic_validation"
    ]

    == "FAIL"

).sum()


total_queries = len(
    combined_df
)


passed_queries = (

    combined_df[
        "automatic_validation"
    ]

    == "PASS"

).sum()


failed_queries = (

    combined_df[
        "automatic_validation"
    ]

    == "FAIL"

).sum()


if total_context_items > 0:

    item_pass_rate = (

        passed_items

        /

        total_context_items

        *

        100

    )

else:

    item_pass_rate = 0


if total_queries > 0:

    query_pass_rate = (

        passed_queries

        /

        total_queries

        *

        100

    )

else:

    query_pass_rate = 0


# ============================================================
# SUMMARY REPORT
# ============================================================

summary_data = [

    [

        "Total Queries",

        total_queries

    ],

    [

        "Passed Queries",

        passed_queries

    ],

    [

        "Failed Queries",

        failed_queries

    ],

    [

        "Query Pass Rate (%)",

        round(

            query_pass_rate,

            2

        )

    ],

    [

        "Total Context Items",

        total_context_items

    ],

    [

        "Passed Context Items",

        passed_items

    ],

    [

        "Failed Context Items",

        failed_items

    ],

    [

        "Context Item Pass Rate (%)",

        round(

            item_pass_rate,

            2

        )

    ],

    [

        "Queries With Context",

        (

            combined_df[
                "context_present"
            ]

            .sum()

            if not combined_df.empty

            else 0

        )

    ],

    [

        "Empty Contexts",

        (

            (

                ~combined_df[
                    "context_present"
                ]

            )

            .sum()

            if not combined_df.empty

            else 0

        )

    ],

    [

        "Duplicate Chunk IDs",

        (

            validation_df[
                "duplicate_chunk_id"
            ]

            .sum()

            if not validation_df.empty

            else 0

        )

    ],

    [

        "Invalid Content Types",

        (

            (

                ~validation_df[
                    "content_type_valid"
                ]

            )

            .sum()

            if not validation_df.empty

            else 0

        )

    ],

    [

        "Invalid Rank Sequences",

        (

            (

                ~validation_df[
                    "rank_sequence_valid"
                ]

            )

            .sum()

            if not validation_df.empty

            else 0

        )

    ]

]


summary_report_df = pd.DataFrame(

    summary_data,

    columns=[

        "Validation Metric",

        "Value"

    ]

)


# ============================================================
# CONTENT TYPE SUMMARY
# ============================================================

if (
    not validation_df.empty
    and
    items_type_col is not None
):

    content_type_summary = (

        validation_df[

            items_type_col

        ]

        .astype(str)

        .str.upper()

        .value_counts()

        .reset_index()

    )


    content_type_summary.columns = [

        "Content Type",

        "Count"

    ]

else:

    content_type_summary = pd.DataFrame(

        columns=[

            "Content Type",

            "Count"

        ]

    )


# ============================================================
# FAILED CONTEXT ITEMS
# ============================================================

if not validation_df.empty:

    failed_items_df = validation_df[

        validation_df[
            "automatic_validation"
        ]

        == "FAIL"

    ].copy()

else:

    failed_items_df = pd.DataFrame()


# ============================================================
# FAILED COMBINED CONTEXT
# ============================================================

if not combined_df.empty:

    failed_queries_df = combined_df[

        combined_df[
            "automatic_validation"
        ]

        == "FAIL"

    ].copy()

else:

    failed_queries_df = pd.DataFrame()


# ============================================================
# RECOMMENDATIONS
# ============================================================

recommendations = []


if failed_items > 0:

    recommendations.append(

        [

            "Context Items",

            "Some context items failed validation. Review the Context Item Validation sheet."

        ]

    )


if failed_queries > 0:

    recommendations.append(

        [

            "Combined Context",

            "Some queries have invalid or empty combined context."

        ]

    )


if (
    not validation_df.empty
    and
    validation_df[
        "duplicate_chunk_id"
    ].sum()
    > 0
):

    recommendations.append(

        [

            "Duplicate Chunks",

            "Duplicate chunk IDs were found in the built context."

        ]

    )


if (
    not combined_df.empty
    and
    (
        ~combined_df[
            "top_k_valid"
        ]
    ).sum()
    > 0
):

    recommendations.append(

        [

            "Top-K",

            "Some queries contain more than the configured Top-K context items."

        ]

    )


if (
    not validation_df.empty
    and
    (
        ~validation_df[
            "reranker_score_present"
        ]
    ).sum()
    > 0
):

    recommendations.append(

        [

            "Reranker Scores",

            "Some context items do not contain reranker scores."

        ]

    )


if len(
    recommendations
) == 0:

    recommendations.append(

        [

            "Overall Status",

            "All automatic context validation checks passed successfully."

        ]

    )


recommendations_df = pd.DataFrame(

    recommendations,

    columns=[

        "Category",

        "Recommendation"

    ]

)


# ============================================================
# CREATE OUTPUT DIRECTORY
# ============================================================

os.makedirs(

    OUTPUT_DIR,

    exist_ok=True

)


# ============================================================
# SAVE VALIDATION REPORT
# ============================================================

print()
print("=" * 70)
print("CREATING CONTEXT VALIDATION REPORT")
print("=" * 70)


try:

    with pd.ExcelWriter(

        OUTPUT_FILE,

        engine="openpyxl"

    ) as writer:


        # Sheet 1

        summary_report_df.to_excel(

            writer,

            sheet_name="Summary",

            index=False

        )


        # Sheet 2

        sheet_validation_df.to_excel(

            writer,

            sheet_name="Sheet Validation",

            index=False

        )


        # Sheet 3

        combined_df.to_excel(

            writer,

            sheet_name="Combined Context Validation",

            index=False

        )


        # Sheet 4

        validation_df.to_excel(

            writer,

            sheet_name="Context Item Validation",

            index=False

        )


        # Sheet 5

        content_type_summary.to_excel(

            writer,

            sheet_name="Content Type Summary",

            index=False

        )


        # Sheet 6

        failed_items_df.to_excel(

            writer,

            sheet_name="Failed Context Items",

            index=False

        )


        # Sheet 7

        failed_queries_df.to_excel(

            writer,

            sheet_name="Failed Queries",

            index=False

        )


        # Sheet 8

        recommendations_df.to_excel(

            writer,

            sheet_name="Recommendations",

            index=False

        )


    print()
    print(
        "✅ Validation report created successfully!"
    )


except Exception as e:

    print()
    print(
        "❌ Error creating validation report!"
    )

    print(
        "Error:",
        e
    )

    raise SystemExit(1)


# ============================================================
# FORMAT EXCEL
# ============================================================

print()
print(
    "Formatting Excel report..."
)


try:

    workbook = load_workbook(

        OUTPUT_FILE

    )


    for worksheet in workbook.worksheets:


        # Freeze first row

        worksheet.freeze_panes = "A2"


        # Wrap text

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = (

                    cell.alignment.copy(

                        wrap_text=True

                    )

                )


        # Auto-adjust column width

        for column_cells in worksheet.columns:

            max_length = 0


            column_letter = (

                column_cells[
                    0
                ]

                .column_letter

            )


            for cell in column_cells:

                try:

                    cell_length = len(

                        str(
                            cell.value
                        )

                    )


                    if cell_length > max_length:

                        max_length = cell_length


                except:

                    pass


            worksheet.column_dimensions[
                column_letter
            ].width = min(

                max_length + 2,

                60

            )


    workbook.save(

        OUTPUT_FILE

    )


    print(
        "✅ Excel formatting completed!"
    )


except Exception as e:

    print(
        "⚠️ Excel formatting warning:"
    )

    print(
        e
    )


# ============================================================
# FINAL OUTPUT
# ============================================================

print()
print("=" * 70)
print("CONTEXT VALIDATION COMPLETED")
print("=" * 70)

print()

print(
    "Total queries:",
    total_queries
)

print(
    "Passed queries:",
    passed_queries
)

print(
    "Failed queries:",
    failed_queries
)

print()

print(
    "Total context items:",
    total_context_items
)

print(
    "Passed context items:",
    passed_items
)

print(
    "Failed context items:",
    failed_items
)

print()

print(
    "Query pass rate:",
    round(

        query_pass_rate,

        2

    ),

    "%"

)

print(
    "Context item pass rate:",
    round(

        item_pass_rate,

        2

    ),

    "%"

)

print()

print(
    "Validation report saved at:"
)

print(
    OUTPUT_FILE
)

print()
print("=" * 70)
print("NEXT STEP")
print("=" * 70)

print()

print(
    "Review the Excel validation report."
)

print()

print(
    "If validation is successful,"
)

print(
    "the next stage is LLM / RAG Answer Generation."
)

print("=" * 70)